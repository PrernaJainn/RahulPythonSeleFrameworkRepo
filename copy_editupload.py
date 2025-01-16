import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


def update_excel_data(filePath, searchTerm, colName, new_value):
    # Open the Excel file
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    # Find the column for the given column name (e.g., "price")
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    # Find the row where the given search term (fruit name) is located
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    # If row and column are found, update the cell with the new value
    if "row" in Dict and "col" in Dict:
        sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
        book.save(filePath)  # Save the file after updating
    else:
        print("Error: Row or Column not found.")


file_path = 'C:\\Users\\prerna.jain\\Downloads\\download.xlsx'
fruit_name = "Mango"
newValue = "299"

# Start the WebDriver
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

# Click the download button to download the file
driver.find_element(By.ID, "downloadButton").click()

# Update the Excel file with the new price value for the fruit
update_excel_data(file_path, fruit_name, "price", newValue)

# Upload the updated file
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

# Wait for the success toast message after uploading
wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))

# Print the success message from the toast
print(driver.find_element(*toast_locator).text)

# Get the column ID for the "Price" column
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")

# Get the actual price of the fruit
actual_price = driver.find_element(By.XPATH, f"//div[text()='{fruit_name}']/parent::div/parent::div/div[@id='cell-{priceColumn}-undefined']").text

# Clean up the actual price (remove extra spaces, if any)
actual_price = actual_price.strip()
time.sleep(2)

# Print the expected vs actual price for debugging
print(f"Expected Price: {newValue}, Actual Price: {actual_price}")

# Check if the price was updated correctly
assert actual_price == newValue, f"Price mismatch! Expected {newValue} but got {actual_price}"
print(f"Price of {fruit_name} is updated to {actual_price}")

# Close the WebDriver
driver.quit()
