# import openpyxl
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# def update_excel_data(filePath, searchTerm, colName, new_value):   # 4 parameters passes here
#     book=openpyxl.load_workbook(filePath)   #book = openpyxl.load_workbook("C:\\Users\\prerna.jain\\Documents\\PythonDemoExcel.xlsx")
#     sheet=book.active                         # file location using dynamically and active sheet should select
#     Dict = {}
#
#     # Find column matching colName (e.g., "price")
#     for i in range(1,sheet.max_column + 1):
#         if sheet.cell(row = 1,column=i).value == colName:            #when i = 4 it matches,dynamically using the value of fruit price
#             Dict["col"]=i                                                #if sheet.cell(row = 1 , column= i).value =="price":
#
#
#     # Find row matching searchTerm (fruit name)
#     for i in range(1, sheet.max_row + 1):
#         for j in range(1, sheet.max_column + 1):               #outerloop for rows and inner loop for columns,row = 3,column=1
#             if sheet.cell(row=i, column=j).value == searchTerm:  #if sheet.cell( row = i , column= j).value =="Apple" :
#                 Dict["row"]= i                                                       # dynamically using the value of fruit
#                 break
#
#     # Debugging to check if correct row and column are found
#     print(f"Found Row: {Dict.get('row')}, Found Column: {Dict.get('col')}")
#
#
#     # Check if row and column are set, then update cell
#     if "row" in Dict and "col" in Dict:                                  #sheet.cell(row=Dict["row"], column=Dict["col"]).value="500"
#       sheet.cell(row=Dict["row"],column=Dict["col"]).value = new_value         #dynamically using the value
#       book.save(file_path)                                               # after updating the value save the Excel
#     else:
#         print("Error: Row or Column not found.")
#
#
# file_path ="'C:\\Users\\prerna.jain\\Downloads\\download.xlsx'"                #file path store in a variable
# fruit_name = "Apple"
# newValue ="999"
# driver = webdriver.Chrome()
# driver.implicitly_wait(5)
# driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
# driver.find_element(By.ID, "downloadButton").click()
#
# # edit the Excel with the updated values (type = file should show in css to identify)
# update_excel_data(file_path, fruit_name, "price", newValue)        #calling the methods for these arguments
#
# #upload
# file_input = driver.find_element(By.CSS_SELECTOR, " input[type='file']")
# file_input.send_keys(file_path)
# wait = WebDriverWait(driver, 5)
# toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
#
# # Ensure the upload is complete before checking the price
# wait.until(EC.visibility_of_element_located(toast_locator))   # Wait for the toast message to disappear
# #time.sleep(2)        #Sleep for a few seconds to allow the page to refresh (optional)
#
#
# print(driver.find_element(*toast_locator).text)
# priceColumn=driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
# #actual_price = driver.find_element(By.XPATH, "//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + str(priceColumn) + "-undefined']").text
# actual_price = driver.find_element(By.XPATH, f"//div[text()='{fruit_name}']/parent::div/parent::div/div[@id='cell-{priceColumn}-undefined']").text
#
# #Print the actual price for debugging
# print(f"Expected Price: {newValue}, Actual Price: {actual_price}")
# # Check if the price is updated correctly
# assert actual_price == newValue, f"Price mismatch! Expected {newValue} but got {actual_price}"
# print(f"Price of {fruit_name} is updated to {actual_price}")
# #assert actual_price == newValue                                               #print(actual_price ,"Price of the fruit is :")
#
#
# # from selenium.webdriver.chrome import webdriver
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl


def update_excel_data(filePath, searchTerm, colName, new_value):
    book=openpyxl.load_workbook(filePath)
    sheet=book.active
    Dict={}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"]=i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"]=i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value=new_value
    book.save(file_path)


file_path='C:\\Users\\prerna.jain\\Downloads\\download.xlsx'
fruit_name="Apple"
newValue="980"
driver=webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

# edit the excel with updated value
update_excel_data(file_path, fruit_name, "price", newValue)

# upload
file_input=driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait=WebDriverWait(driver, 5)
toast_locator=(By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
priceColumn=driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price=driver.find_element(By.XPATH,
                                 "//div[text()='" + fruit_name + "']/parent::div/parent::div/div[@id='cell-" + priceColumn + "-undefined']").text
assert actual_price == newValue

