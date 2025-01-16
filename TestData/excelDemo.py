#read the value from the sheet
import openpyxl
book =openpyxl.load_workbook("C:\\Users\\prerna.jain\\Documents\\PythonDemoExcel.xlsx") # file location
sheet =book.active   #active sheet should select
Dict ={}                              #empty dictionary

cell =sheet.cell(row=1,column=2)   #b1
cell.value
print( "Cell Value is : " , cell.value)

# write the value in the sheet (b2)
sheet.cell(row=2, column=2).value = "Prerna"
print(sheet.cell(row=2,column=2).value)

#how many maximum rows /columns present in the sheet (sheet level)
print(sheet.max_row)
print(sheet.max_column)

#if you want the direct the name of the sheet row and colum name use this shortcut
print(sheet['A5'].value)

#print all the values in the sheet using for loop
#rowfirst = 1  column = 1  #name  updating only row not a column to overcome this issue use inner column forloop
for rowfirst in range(1,sheet.max_row+1):   #to get rows
    # I need only testcases2 value
    if sheet.cell(row=rowfirst , column=1).value == "Testcase2":

         for secondcolumn in range(2,sheet.max_column+1):   # to get columns ,if start with it skip the testcases name 2 shows only value
          # print(sheet.cell(row= rowfirst, column= secondcolumn).value) rowfirst = 2  column = ----4  name
           Dict[sheet.cell(row=1, column= secondcolumn).value]=sheet.cell(row= rowfirst, column= secondcolumn).value
           # Dict ["lastname "]= "Johnson"

print(Dict)        # value captured from the Excel sheet